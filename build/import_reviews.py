"""Merge Outscraper Google review exports into data/reviews.json.

Usage:  python3 build/import_reviews.py <file.xlsx|file.csv> [more files...]

Maps each row back to a city slug by place_id via data/harvest-queue.json, so
a review can never be attributed to the wrong city. Reviewer names are cut to
"First L." to match what is already on file: the reviews are public on Google,
but there is no reason to republish full names.
"""
import json, os, re, sys, csv

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REVIEWS = os.path.join(ROOT, "data/reviews.json")
QUEUE = os.path.join(ROOT, "data/harvest-queue.json")


def slug_map():
    q = json.load(open(QUEUE))["queue"]
    return {x["place_id"]: (x["slug"], x["city"]) for x in q}


def clean(t):
    t = re.sub(r'<br\s*/?>', ' ', str(t or ""))
    t = re.sub(r'<[^>]+>', ' ', t)
    t = t.replace(" ", " ").replace("&amp;", "&")
    t = t.replace("\u2014", ", ")                    # house rule: no em dashes
    return re.sub(r'\s+', ' ', t).strip()


def short_name(n):
    parts = clean(n).split()
    if not parts: return "A Google user"
    if len(parts) == 1: return parts[0]
    return f"{parts[0]} {parts[1][0]}."


def rows_from(path):
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        ws = openpyxl.load_workbook(path, read_only=True)[openpyxl.load_workbook(path, read_only=True).sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        for r in it: yield dict(zip(hdr, r))
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f): yield row


def main(paths):
    smap = slug_map()
    existing = json.load(open(REVIEWS))
    seen = {clean(x["quote"])[:80] for x in existing}
    added = skipped = unmapped = 0
    for p in paths:
        for r in rows_from(p):
            pid = r.get("place_id")
            if pid not in smap:
                unmapped += 1; continue
            text = clean(r.get("review_text"))
            if len(text) < 40:                      # star-only, nothing to quote
                skipped += 1; continue
            if text[:80] in seen:
                skipped += 1; continue
            seen.add(text[:80])
            slug, city = smap[pid]
            try: rating = int(float(r.get("review_rating") or 5))
            except Exception: rating = 5
            rec = {"franchise": "TX", "name": short_name(r.get("author_title")),
                   "city": city.split(" (")[0], "state": "TX", "rating": rating,
                   "quote": text, "source": "google", "slug": slug,
                   "date": str(r.get("review_datetime_utc") or "")[:10]}
            ans = clean(r.get("owner_answer"))
            if ans: rec["owner_reply"] = ans
            existing.append(rec); added += 1
    json.dump(existing, open(REVIEWS, "w"), indent=1)
    import collections
    c = collections.Counter(x["slug"] for x in existing)
    print(f"added {added}, skipped {skipped} (dupes/too short), unmapped place_ids {unmapped}")
    print(f"reviews.json now {len(existing)} across {len(c)} cities")
    for s, n in c.most_common(): print(f"   {n:>3}  {s}")


if __name__ == "__main__":
    main(sys.argv[1:] or sys.exit("usage: import_reviews.py <export files...>"))
